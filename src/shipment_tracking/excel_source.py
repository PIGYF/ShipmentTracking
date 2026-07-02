from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class ShipmentInput:
    row_number: int
    tracking_number: str
    forwarder: str


def read_shipments_from_excel(path: str | Path, sheet_name: str | None = None) -> list[ShipmentInput]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    header_row, headers = _find_header_row(sheet)

    tracking_idx = _find_column(
        headers,
        [
            "\u8fd0\u5355",
            "\u63d0\u5355\u53f7",
            "bl number",
            "b/l",
            "house bill",
            "master bill",
            "mbl",
            "hbl",
            "tracking_number",
        ],
    )
    forwarder_idx = _find_column(
        headers,
        ["\u8d27\u4ee3", "forwarder", "carrier", "shipping line", "agent"],
    )

    rows: list[ShipmentInput] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        tracking_value = row[tracking_idx] if tracking_idx < len(row) else None
        forwarder_value = row[forwarder_idx] if forwarder_idx < len(row) else None
        if tracking_value is None:
            continue
        rows.append(
            ShipmentInput(
                row_number=row_number,
                tracking_number=str(tracking_value).strip(),
                forwarder=str(forwarder_value or "").strip(),
            )
        )
    return rows


def _find_header_row(sheet) -> tuple[int, list[str]]:
    aliases = {
        "\u8fd0\u5355",
        "\u63d0\u5355\u53f7",
        "\u8d27\u4ee3",
        "bl number",
        "b/l",
        "house bill",
        "forwarder",
        "carrier",
    }
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        headers = [str(value).strip() if value is not None else "" for value in row]
        lowered = [header.lower() for header in headers]
        if any(alias in lowered for alias in aliases):
            return row_number, headers
    raise ValueError("Could not find a shipment header row in the first 20 rows.")


def _find_column(headers: list[str], names: list[str]) -> int:
    lowered = [header.lower() for header in headers]
    for name in names:
        needle = name.lower()
        for index, header in enumerate(lowered):
            if header == needle or needle in header:
                return index
    raise ValueError(f"Missing required column. Expected one of: {', '.join(names)}")
