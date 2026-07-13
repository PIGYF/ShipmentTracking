from __future__ import annotations

from datetime import datetime
import gc
import json
import locale
import os
from pathlib import Path
import re
import subprocess
import tempfile
import time
from typing import Iterable
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .models import TrackingRecord
from .time_utils import date_only


COL_BILL_NO = "\u63d0\u5355\u53f7"
COL_FORWARDER = "\u8d27\u4ee3"
COL_STATUS = "\u72b6\u6001\u663e\u793a"
COL_CALL_FOR_PICKUP = "\u901a\u77e5\u63d0\u8d27\u65e5\u671f"
COL_ACTUAL_PICKUP = "\u5b9e\u9645\u63d0\u8d27\u65e5\u671f"
COL_DEPARTURE = "\u79bb\u6e2f\u65e5\u671f"
COL_ARRIVAL = "\u5230\u6e2f\u65e5\u671f"
COL_EXCEPTION = "\u5f02\u5e38\u8bb0\u5f55"
COL_TRACKING_NOTE = "\u8ffd\u8e2a\u8bb0\u5f55"

COL_BILL_NO_EN = "BILL_NO"
COL_FORWARDER_EN = "GOODS_YARD"
COL_STATUS_EN = "DISPLAY_STATUS"
COL_CALL_FOR_PICKUP_EN = "CALL FOR PICK UP DATE"
COL_ACTUAL_PICKUP_EN = "ACTUAL PICK UP DATE"
COL_DEPARTURE_EN = "DEPARTURE_DATE"
COL_ARRIVAL_EN = "ARRIVAL_DATE"

FILL_CHANGED = PatternFill("solid", fgColor="FFF2CC")
FILL_ERROR = PatternFill("solid", fgColor="F4CCCC")
FILL_WARNING = PatternFill("solid", fgColor="FCE5CD")
DATED_REMARK_PREFIX = re.compile(r"^\[\d{4}-\d{2}-\d{2} \d{2}:\d{2}(?::\d{2})?\]\s*")


def update_tracking_workbook(
    source_path: str | Path,
    records: Iterable[TrackingRecord],
    output_path: str | Path,
    sheet_name: str | None = None,
    remarks: dict[str, str] | None = None,
    run_label: str | None = None,
) -> int:
    records = list(records)
    remarks = remarks or {}
    if _should_use_excel_com():
        return _update_tracking_workbook_with_excel_com(source_path, records, output_path, sheet_name, remarks, run_label)
    return _update_tracking_workbook_with_openpyxl(source_path, records, output_path, sheet_name, remarks, run_label)


def _should_use_excel_com() -> bool:
    return os.name == "nt" and os.getenv("SHIPMENT_TRACKING_USE_EXCEL_COM", "1") != "0"


def _update_tracking_workbook_with_openpyxl(
    source_path: str | Path,
    records: list[TrackingRecord],
    output_path: str | Path,
    sheet_name: str | None,
    remarks: dict[str, str],
    run_label: str | None,
) -> int:
    source = Path(source_path)
    output = Path(output_path)
    preserve_source = source
    temp_preserve: Path | None = None
    if source.resolve() == output.resolve():
        with tempfile.NamedTemporaryFile(delete=False, suffix=source.suffix) as temp_file:
            temp_file.write(source.read_bytes())
            temp_preserve = Path(temp_file.name)
            preserve_source = temp_preserve

    workbook = load_workbook(preserve_source)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    header_row, headers = _find_header_row(sheet)
    columns = _column_map(headers)

    key_col = _require_column(columns, [COL_BILL_NO, COL_BILL_NO_EN, "\u8fd0\u5355", "bl number", "b/l", "house bill", "tracking_number"])
    carrier_col = _find_column(columns, [COL_FORWARDER, COL_FORWARDER_EN, "forwarder", "carrier"])
    call_for_pickup_col = _require_column(columns, [COL_CALL_FOR_PICKUP, COL_CALL_FOR_PICKUP_EN])
    actual_pickup_col = _require_column(columns, [COL_ACTUAL_PICKUP, COL_ACTUAL_PICKUP_EN])
    departure_col = _require_column(columns, [COL_DEPARTURE, COL_DEPARTURE_EN])
    arrival_col = _require_column(columns, [COL_ARRIVAL, COL_ARRIVAL_EN])
    tracking_note_col = _require_column(columns, [COL_TRACKING_NOTE])

    record_map = {record.tracking_number: record for record in records}
    remarks = remarks or {}
    updated = 0

    for row in range(header_row + 1, sheet.max_row + 1):
        tracking_number = str(sheet.cell(row=row, column=key_col).value or "").strip()
        record = record_map.get(tracking_number)
        remark = remarks.get(tracking_number)
        if not record and not remark:
            continue

        if record and carrier_col:
            carrier = str(sheet.cell(row=row, column=carrier_col).value or "").strip().upper()
            if record.carrier.upper() not in carrier:
                continue

        row_changed = False
        if record and record.found:
            row_changed, remark = _update_record_date_cell(
                sheet,
                row,
                call_for_pickup_col,
                record.call_for_pickup_date,
                record,
                "CALL_FOR_PICKUP_DATE",
                run_label,
                remark,
                row_changed,
            )
            row_changed, remark = _update_record_date_cell(
                sheet,
                row,
                actual_pickup_col,
                record.actual_pickup,
                record,
                "ACTUAL_PICKUP_DATE",
                run_label,
                remark,
                row_changed,
            )
            row_changed, remark = _update_record_date_cell(
                sheet,
                row,
                departure_col,
                record.departure_date,
                record,
                "DEPARTURE_DATE",
                run_label,
                remark,
                row_changed,
            )
            if record.arrival_date:
                changed, old_display, new_display = _set_date_if_changed(
                    sheet.cell(row=row, column=arrival_col),
                    date_only(record.arrival_date),
                    date_format="m/d/yyyy",
                )
                row_changed |= changed
                if changed:
                    change_remark = _date_change_remark(
                        run_label,
                        record.carrier,
                        record.arrival_date_type,
                        old_display,
                        new_display,
                    )
                    remark = _join_remarks(remark, change_remark)
                elif _is_actual_arrival(record.arrival_date_type):
                    confirm_remark = _actual_arrival_confirmation_remark(
                        run_label,
                        record.carrier,
                        new_display,
                    )
                    remark = _join_remarks(remark, confirm_remark)

        if remark:
            cell = sheet.cell(row=row, column=tracking_note_col)
            cell.value = _merge_tracking_note(cell.value, remark)
            if remark.startswith("ERROR") or remark.startswith("NOT_FOUND"):
                cell.fill = FILL_ERROR
            elif "NO_ARRIVAL_DATE" in remark:
                cell.fill = FILL_WARNING

        if row_changed or remark:
            updated += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    gc.collect()
    _restore_unmanaged_package_parts(preserve_source, output)
    if temp_preserve and temp_preserve.exists():
        _unlink_with_retry(temp_preserve, raise_on_failure=False)
    return updated


def _update_tracking_workbook_with_excel_com(
    source_path: str | Path,
    records: list[TrackingRecord],
    output_path: str | Path,
    sheet_name: str | None,
    remarks: dict[str, str],
    run_label: str | None,
) -> int:
    source = Path(source_path).resolve()
    output = Path(output_path).resolve()
    payload = _excel_com_payload(records, remarks)

    with tempfile.TemporaryDirectory(prefix="shipment_tracking_excel_") as temp_dir:
        temp_path = Path(temp_dir)
        payload_path = temp_path / "payload.json"
        script_path = temp_path / "update_excel.ps1"
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        script_path.write_text(_EXCEL_COM_SCRIPT, encoding="utf-8-sig")

        completed = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                str(source),
                str(output),
                sheet_name or "",
                str(payload_path),
                run_label or "",
            ],
            capture_output=True,
        )

    stdout = _decode_subprocess_output(completed.stdout)
    stderr = _decode_subprocess_output(completed.stderr)
    if completed.returncode != 0:
        details = _subprocess_details(stdout, stderr)
        raise RuntimeError(f"Excel COM update failed: {details}")

    for line in stdout.splitlines():
        if line.startswith("UPDATED_ROWS="):
            return int(line.split("=", 1)[1])
    raise RuntimeError(f"Excel COM update did not report updated rows: {_subprocess_details(stdout, stderr)}")


def _excel_com_payload(records: list[TrackingRecord], remarks: dict[str, str]) -> list[dict[str, str | bool | None]]:
    record_map = {record.tracking_number: record for record in records}
    tracking_numbers = sorted(set(record_map) | set(remarks))
    payload: list[dict[str, str | bool | None]] = []
    for tracking_number in tracking_numbers:
        record = record_map.get(tracking_number)
        payload.append(
            {
                "tracking_number": tracking_number,
                "carrier": record.carrier if record else None,
                "found": record.found if record else False,
                "arrival_date": _display_date(record.arrival_date) if record and record.arrival_date else None,
                "arrival_date_type": record.arrival_date_type if record else None,
                "call_for_pickup_date": _display_date(record.call_for_pickup_date) if record and record.call_for_pickup_date else None,
                "actual_pickup": _display_date(record.actual_pickup) if record and record.actual_pickup else None,
                "departure_date": _display_date(record.departure_date) if record and record.departure_date else None,
                "remark": remarks.get(tracking_number),
            }
        )
    return payload


def _display_date(value: datetime) -> str:
    return f"{value.month}/{value.day}/{value.year}"


def _decode_subprocess_output(value: bytes | None) -> str:
    if not value:
        return ""
    for encoding in ("utf-8", locale.getpreferredencoding(False)):
        try:
            return value.decode(encoding)
        except UnicodeDecodeError:
            continue
    return value.decode("utf-8", errors="replace")


def _subprocess_details(stdout: str, stderr: str) -> str:
    parts = []
    if stderr.strip():
        parts.append(f"stderr: {stderr.strip()}")
    if stdout.strip():
        parts.append(f"stdout: {stdout.strip()}")
    return "\n".join(parts) if parts else "(no output)"


def _restore_unmanaged_package_parts(source_path: Path, output_path: Path) -> None:
    if source_path.resolve() == output_path.resolve():
        original_bytes = source_path.read_bytes()
        with tempfile.NamedTemporaryFile(delete=False, suffix=source_path.suffix) as temp_original:
            temp_original.write(original_bytes)
            temp_original_path = Path(temp_original.name)
    else:
        temp_original_path = source_path

    temp_output = output_path.with_suffix(f"{output_path.suffix}.tmp")
    try:
        with ZipFile(temp_original_path) as source_zip, ZipFile(output_path) as output_zip:
            source_names = set(source_zip.namelist())
            output_names = set(output_zip.namelist())
            preserved_names = {
                name
                for name in source_names
                if name.startswith("customXml/") or name.startswith("[trash]/")
            }

            with ZipFile(temp_output, "w", ZIP_DEFLATED) as merged_zip:
                for item in output_zip.infolist():
                    data = output_zip.read(item.filename)
                    if item.filename == "[Content_Types].xml" and item.filename in source_names:
                        data = _merge_content_types(source_zip.read(item.filename), data)
                    elif item.filename == "xl/_rels/workbook.xml.rels" and item.filename in source_names:
                        data = _merge_workbook_relationships(source_zip.read(item.filename), data)
                    merged_zip.writestr(item, data)

                for name in sorted(preserved_names - output_names):
                    merged_zip.writestr(name, source_zip.read(name))

        _replace_with_retry(temp_output, output_path)
    finally:
        if source_path.resolve() == output_path.resolve() and temp_original_path.exists():
            temp_original_path.unlink()
        if temp_output.exists():
            temp_output.unlink()


def _merge_content_types(source_xml: bytes, output_xml: bytes) -> bytes:
    namespace = "http://schemas.openxmlformats.org/package/2006/content-types"
    ET.register_namespace("", namespace)
    source_root = ET.fromstring(source_xml)
    output_root = ET.fromstring(output_xml)
    existing_defaults = {element.attrib.get("Extension") for element in output_root.findall(f"{{{namespace}}}Default")}
    existing_overrides = {element.attrib.get("PartName") for element in output_root.findall(f"{{{namespace}}}Override")}

    for element in source_root:
        if element.tag == f"{{{namespace}}}Default":
            key = element.attrib.get("Extension")
            if key not in existing_defaults:
                output_root.append(element)
                existing_defaults.add(key)
        elif element.tag == f"{{{namespace}}}Override":
            key = element.attrib.get("PartName")
            if key and (key.startswith("/customXml/") or key.startswith("/[trash]/")) and key not in existing_overrides:
                output_root.append(element)
                existing_overrides.add(key)

    return ET.tostring(output_root, encoding="utf-8", xml_declaration=True)


def _replace_with_retry(source: Path, target: Path) -> None:
    last_error: PermissionError | None = None
    for _ in range(10):
        try:
            _unlink_with_retry(target)
            source.replace(target)
            return
        except PermissionError as exc:
            last_error = exc
            time.sleep(0.5)
    if last_error:
        raise last_error


def _unlink_with_retry(path: Path, *, raise_on_failure: bool = True) -> None:
    if not path.exists():
        return
    last_error: PermissionError | None = None
    for _ in range(10):
        try:
            path.unlink()
            return
        except PermissionError as exc:
            last_error = exc
            time.sleep(0.5)
    if last_error and raise_on_failure:
        raise last_error


def _merge_workbook_relationships(source_xml: bytes, output_xml: bytes) -> bytes:
    namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    ET.register_namespace("", namespace)
    source_root = ET.fromstring(source_xml)
    output_root = ET.fromstring(output_xml)
    existing_targets = {element.attrib.get("Target") for element in output_root.findall(f"{{{namespace}}}Relationship")}

    for element in source_root.findall(f"{{{namespace}}}Relationship"):
        rel_type = element.attrib.get("Type", "")
        target = element.attrib.get("Target")
        if rel_type.endswith("/customXml") and target not in existing_targets:
            output_root.append(element)
            existing_targets.add(target)

    return ET.tostring(output_root, encoding="utf-8", xml_declaration=True)


def _set_if_changed(cell, new_value, date_format: str | None = None) -> bool:
    if new_value is None:
        return False
    old_value = _normalize_cell_value(cell.value)
    comparable = _normalize_cell_value(new_value)
    if old_value == comparable:
        return False
    cell.value = new_value
    if date_format:
        cell.number_format = date_format
    cell.fill = FILL_CHANGED
    return True


def _set_date_if_changed(cell, new_value, date_format: str | None = None) -> tuple[bool, str, str]:
    if new_value is None:
        return False, "", ""
    old_value = _normalize_cell_value(cell.value)
    comparable = _normalize_cell_value(new_value)
    if old_value == comparable:
        return False, _display_cell_value(cell.value), _display_cell_value(new_value)
    old_display = _display_cell_value(cell.value)
    new_display = _display_cell_value(new_value)
    cell.value = new_value
    if date_format:
        cell.number_format = date_format
    cell.fill = FILL_CHANGED
    return True, old_display, new_display


def _update_record_date_cell(
    sheet,
    row: int,
    column: int,
    value: datetime | None,
    record: TrackingRecord,
    field_label: str,
    run_label: str | None,
    remark: str | None,
    row_changed: bool,
) -> tuple[bool, str | None]:
    if not value:
        return row_changed, remark
    changed, old_display, new_display = _set_date_if_changed(
        sheet.cell(row=row, column=column),
        date_only(value),
        date_format="m/d/yyyy",
    )
    if changed:
        change_remark = _date_field_change_remark(
            run_label,
            record.carrier,
            field_label,
            old_display,
            new_display,
        )
        remark = _join_remarks(remark, change_remark)
        row_changed = True
    return row_changed, remark


def _normalize_cell_value(value):
    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(value, str):
        text = value.strip()
        for date_format in ("%m/%d/%Y", "%-m/%-d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, date_format)
            except ValueError:
                continue
    return value


def _merge_tracking_note(existing, addition: str) -> str:
    current_lines = _remark_lines(str(existing)) if existing is not None else []
    for addition_line in _remark_lines(addition):
        addition_body = _remark_body(addition_line)
        matching_indexes = [
            index
            for index, current_line in enumerate(current_lines)
            if _remark_body(current_line) == addition_body
        ]
        if not matching_indexes:
            current_lines.append(addition_line)
            continue
        current_lines[matching_indexes[0]] = addition_line
        for index in reversed(matching_indexes[1:]):
            del current_lines[index]
    return "\n".join(current_lines)


def _remark_lines(value: str) -> list[str]:
    return [line.strip() for line in value.splitlines() if line.strip()]


def _remark_body(line: str) -> str:
    return DATED_REMARK_PREFIX.sub("", line.strip(), count=1)


def _join_remarks(existing: str | None, addition: str) -> str:
    return f"{existing}\n{addition}" if existing else addition


def _date_change_remark(
    run_label: str | None,
    carrier: str | None,
    arrival_date_type: str | None,
    old_display: str,
    new_display: str,
) -> str:
    prefix = f"[{run_label}] " if run_label else ""
    carrier_label = f"{carrier}: " if carrier else ""
    type_label = f"{arrival_date_type or 'ARRIVAL'} "
    return f"{prefix}{carrier_label}{type_label}ARRIVAL_DATE_CHANGED: {old_display or '(blank)'} -> {new_display}"


def _date_field_change_remark(
    run_label: str | None,
    carrier: str | None,
    field_label: str,
    old_display: str,
    new_display: str,
) -> str:
    prefix = f"[{run_label}] " if run_label else ""
    carrier_label = f"{carrier}: " if carrier else ""
    return f"{prefix}{carrier_label}{field_label}_CHANGED: {old_display or '(blank)'} -> {new_display}"


def _actual_arrival_confirmation_remark(run_label: str | None, carrier: str | None, arrival_display: str) -> str:
    prefix = f"[{run_label}] " if run_label else ""
    carrier_label = f"{carrier}: " if carrier else ""
    return f"{prefix}{carrier_label}ACTUAL ARRIVAL_CONFIRMED: {arrival_display}"


def _is_actual_arrival(arrival_date_type: str | None) -> bool:
    return str(arrival_date_type or "").upper() == "ACTUAL"


def _display_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return _display_date(value)
    return str(value).strip()


def _find_header_row(sheet) -> tuple[int, list[str]]:
    for required_set in ({COL_BILL_NO, COL_FORWARDER, COL_STATUS}, {COL_BILL_NO_EN, COL_FORWARDER_EN, COL_STATUS_EN}):
        for row_number in range(1, min(sheet.max_row, 20) + 1):
            headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[row_number]]
            lowered = [header.lower() for header in headers]
            if all(required.lower() in lowered for required in required_set):
                return row_number, headers
    raise ValueError("Could not find a shipment header row in the first 20 rows.")


def _column_map(headers: list[str]) -> dict[str, int]:
    return {header.lower(): index + 1 for index, header in enumerate(headers) if header}


def _find_column(columns: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        needle = name.lower()
        for header, column in columns.items():
            if header == needle or needle in header:
                return column
    return None


def _require_column(columns: dict[str, int], names: list[str]) -> int:
    column = _find_column(columns, names)
    if not column:
        raise ValueError(f"Missing required column. Expected one of: {', '.join(names)}")
    return column


_EXCEL_COM_SCRIPT = r'''
param(
    [Parameter(Mandatory=$true)][string]$WorkbookPath,
    [Parameter(Mandatory=$true)][string]$OutputPath,
    [Parameter(Mandatory=$false)][string]$SheetName,
    [Parameter(Mandatory=$true)][string]$PayloadPath,
    [Parameter(Mandatory=$false)][string]$RunLabel
)

$ErrorActionPreference = 'Stop'
[Console]::OutputEncoding = [System.Text.UTF8Encoding]::new($false)
$OutputEncoding = [Console]::OutputEncoding
$payload = Get-Content -Path $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$items = @{}
foreach ($item in $payload) {
    $items[[string]$item.tracking_number] = $item
}

function Normalize([object]$value) {
    if ($null -eq $value) { return '' }
    return ([string]$value).Trim().ToLowerInvariant()
}

function Find-Column($headers, [string[]]$names, [bool]$allowContains) {
    foreach ($name in $names) {
        $needle = Normalize $name
        foreach ($key in $headers.Keys) {
            if ($key -eq $needle) { return $headers[$key] }
        }
    }
    if ($allowContains) {
        foreach ($name in $names) {
            $needle = Normalize $name
            foreach ($key in $headers.Keys) {
                if ($key.Contains($needle)) { return $headers[$key] }
            }
        }
    }
    return $null
}

function Find-Header($sheet, [int]$maxCol) {
    $requiredSets = @(
        @('提单号', '货代', '状态显示'),
        @('BILL_NO', 'GOODS_YARD', 'DISPLAY_STATUS')
    )
    foreach ($requiredSet in $requiredSets) {
        for ($row = 1; $row -le 20; $row++) {
            $headers = @{}
            for ($col = 1; $col -le $maxCol; $col++) {
                $value = $sheet.Cells.Item($row, $col).Text
                if (-not [string]::IsNullOrWhiteSpace($value)) {
                    $headers[(Normalize $value)] = $col
                }
            }
            $ok = $true
            foreach ($required in $requiredSet) {
                if (-not $headers.ContainsKey((Normalize $required))) {
                    $ok = $false
                    break
                }
            }
            if ($ok) {
                return @{ Row = $row; Headers = $headers }
            }
        }
    }
    throw 'Could not find expected header row.'
}

function Display-Date([object]$value) {
    if ($null -eq $value -or [string]::IsNullOrWhiteSpace([string]$value)) { return '' }
    if ($value -is [datetime]) { return $value.ToString('M/d/yyyy', [Globalization.CultureInfo]::InvariantCulture) }
    try {
        if ($value -is [double] -or $value -is [int]) {
            return ([datetime]::FromOADate([double]$value)).ToString('M/d/yyyy', [Globalization.CultureInfo]::InvariantCulture)
        }
    } catch {}
    return ([string]$value).Trim()
}

function Append-Remark([string]$existing, [string]$addition) {
    if ([string]::IsNullOrWhiteSpace($addition)) { return $existing }
    if ([string]::IsNullOrWhiteSpace($existing)) { return $addition }
    return $existing + "`n" + $addition
}

function Remark-Body([string]$line) {
    if ([string]::IsNullOrWhiteSpace($line)) { return '' }
    return ([regex]::Replace($line.Trim(), '^\[\d{4}-\d{2}-\d{2} \d{2}:\d{2}(?::\d{2})?\]\s*', '', 1))
}

function Remark-Lines([string]$value) {
    $lines = [System.Collections.Generic.List[string]]::new()
    if ([string]::IsNullOrWhiteSpace($value)) { return ,$lines }
    foreach ($line in ($value -split "`r?`n")) {
        $trimmed = $line.Trim()
        if (-not [string]::IsNullOrWhiteSpace($trimmed)) {
            $lines.Add($trimmed)
        }
    }
    return ,$lines
}

function Merge-Tracking-Note([string]$existing, [string]$addition) {
    $currentLines = Remark-Lines $existing
    foreach ($additionLine in (Remark-Lines $addition)) {
        $additionBody = Remark-Body $additionLine
        $matchingIndexes = [System.Collections.Generic.List[int]]::new()
        for ($index = 0; $index -lt $currentLines.Count; $index++) {
            if ((Remark-Body $currentLines[$index]) -eq $additionBody) {
                $matchingIndexes.Add($index)
            }
        }

        if ($matchingIndexes.Count -eq 0) {
            $currentLines.Add($additionLine)
            continue
        }

        $currentLines[$matchingIndexes[0]] = $additionLine
        for ($index = $matchingIndexes.Count - 1; $index -ge 1; $index--) {
            $currentLines.RemoveAt($matchingIndexes[$index])
        }
    }
    return [string]::Join("`n", $currentLines)
}

function Update-Date-Cell($cell, [string]$dateText, [string]$carrierLabel, [string]$fieldLabel, [string]$runLabel, [string]$remark) {
    if ([string]::IsNullOrWhiteSpace($dateText)) {
        return @{ Changed = $false; Remark = $remark }
    }
    $newDate = [datetime]::ParseExact($dateText, 'M/d/yyyy', [Globalization.CultureInfo]::InvariantCulture)
    $oldDisplay = Display-Date $cell.Value2
    $newDisplay = $newDate.ToString('M/d/yyyy', [Globalization.CultureInfo]::InvariantCulture)
    if ($oldDisplay -eq $newDisplay) {
        return @{ Changed = $false; Remark = $remark }
    }
    $cell.NumberFormat = 'm/d/yyyy'
    $cell.Value = $newDisplay
    $cell.Interior.Color = 13431551
    $prefix = ''
    if (-not [string]::IsNullOrWhiteSpace($runLabel)) {
        $prefix = '[' + $runLabel + '] '
    }
    if ([string]::IsNullOrWhiteSpace($oldDisplay)) {
        $oldDisplay = '(blank)'
    }
    $newRemark = Append-Remark $remark ($prefix + $carrierLabel + $fieldLabel + '_CHANGED: ' + $oldDisplay + ' -> ' + $newDisplay)
    return @{ Changed = $true; Remark = $newRemark }
}

$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$workbook = $null
$updated = 0

try {
    $workbook = $excel.Workbooks.Open($WorkbookPath)
    if ([string]::IsNullOrWhiteSpace($SheetName)) {
        $sheet = $workbook.Worksheets.Item(1)
    } else {
        $sheet = $null
        foreach ($candidateSheet in $workbook.Worksheets) {
            if ([string]$candidateSheet.Name -eq $SheetName) {
                $sheet = $candidateSheet
                break
            }
        }
        if ($null -eq $sheet) {
            throw "Worksheet not found: $SheetName"
        }
    }

    $used = $sheet.UsedRange
    $maxRow = $used.Row + $used.Rows.Count - 1
    $maxCol = $used.Column + $used.Columns.Count - 1
    $headerInfo = Find-Header $sheet $maxCol
    $headerRow = [int]$headerInfo.Row
    $headers = $headerInfo.Headers

    $keyCol = Find-Column $headers @('提单号', 'BILL_NO', '运单', 'bl number', 'b/l', 'house bill', 'tracking_number') $true
    $carrierCol = Find-Column $headers @('货代', 'GOODS_YARD', 'forwarder', 'carrier') $true
    $callForPickupCol = Find-Column $headers @('通知提货日期', 'CALL FOR PICK UP DATE') $false
    $actualPickupCol = Find-Column $headers @('实际提货日期', 'ACTUAL PICK UP DATE') $false
    $departureCol = Find-Column $headers @('离港日期', 'DEPARTURE_DATE') $false
    $arrivalCol = Find-Column $headers @('到港日期', 'ARRIVAL_DATE') $false
    $trackingNoteCol = Find-Column $headers @('追踪记录') $false

    if ($null -eq $keyCol -or $null -eq $arrivalCol) {
        throw 'Missing required 提单号/BILL_NO or 到港日期/ARRIVAL_DATE column.'
    }
    if ($null -eq $callForPickupCol -or $null -eq $actualPickupCol -or $null -eq $departureCol) {
        throw 'Missing required 通知提货日期/CALL FOR PICK UP DATE, 实际提货日期/ACTUAL PICK UP DATE, or 离港日期/DEPARTURE_DATE column.'
    }
    if ($null -eq $trackingNoteCol) {
        throw 'Missing required 追踪记录 column.'
    }

    for ($row = $headerRow + 1; $row -le $maxRow; $row++) {
        $trackingNumber = ([string]$sheet.Cells.Item($row, $keyCol).Text).Trim()
        if ([string]::IsNullOrWhiteSpace($trackingNumber) -or -not $items.ContainsKey($trackingNumber)) {
            continue
        }
        $item = $items[$trackingNumber]
        if ($null -ne $item.carrier -and $null -ne $carrierCol) {
            $carrier = ([string]$sheet.Cells.Item($row, $carrierCol).Text).Trim().ToUpperInvariant()
            if (-not $carrier.Contains(([string]$item.carrier).ToUpperInvariant())) {
                continue
            }
        }

        $rowChanged = $false
        $remark = [string]$item.remark
        if ($item.found) {
            $carrierLabel = ''
            if (-not [string]::IsNullOrWhiteSpace([string]$item.carrier)) {
                $carrierLabel = [string]$item.carrier + ': '
            }
            $dateResult = Update-Date-Cell $sheet.Cells.Item($row, $callForPickupCol) ([string]$item.call_for_pickup_date) $carrierLabel 'CALL_FOR_PICKUP_DATE' $RunLabel $remark
            $remark = [string]$dateResult.Remark
            if ($dateResult.Changed) { $rowChanged = $true }
            $dateResult = Update-Date-Cell $sheet.Cells.Item($row, $actualPickupCol) ([string]$item.actual_pickup) $carrierLabel 'ACTUAL_PICKUP_DATE' $RunLabel $remark
            $remark = [string]$dateResult.Remark
            if ($dateResult.Changed) { $rowChanged = $true }
            $dateResult = Update-Date-Cell $sheet.Cells.Item($row, $departureCol) ([string]$item.departure_date) $carrierLabel 'DEPARTURE_DATE' $RunLabel $remark
            $remark = [string]$dateResult.Remark
            if ($dateResult.Changed) { $rowChanged = $true }
        }
        if ($item.found -and -not [string]::IsNullOrWhiteSpace([string]$item.arrival_date)) {
            $arrivalCell = $sheet.Cells.Item($row, $arrivalCol)
            $arrivalDate = [datetime]::ParseExact([string]$item.arrival_date, 'M/d/yyyy', [Globalization.CultureInfo]::InvariantCulture)
            $oldDisplay = Display-Date $arrivalCell.Value2
            $newDisplay = $arrivalDate.ToString('M/d/yyyy', [Globalization.CultureInfo]::InvariantCulture)
            $carrierLabel = ''
            if (-not [string]::IsNullOrWhiteSpace([string]$item.carrier)) {
                $carrierLabel = [string]$item.carrier + ': '
            }
            $typeLabel = [string]$item.arrival_date_type
            if ([string]::IsNullOrWhiteSpace($typeLabel)) {
                $typeLabel = 'ARRIVAL'
            }
            if ($oldDisplay -ne $newDisplay) {
                $arrivalCell.NumberFormat = 'm/d/yyyy'
                $arrivalCell.Value = $newDisplay
                $arrivalCell.Interior.Color = 13431551
                $prefix = ''
                if (-not [string]::IsNullOrWhiteSpace($RunLabel)) {
                    $prefix = '[' + $RunLabel + '] '
                }
                if ([string]::IsNullOrWhiteSpace($oldDisplay)) {
                    $oldDisplay = '(blank)'
                }
                $remark = Append-Remark $remark ($prefix + $carrierLabel + $typeLabel + ' ARRIVAL_DATE_CHANGED: ' + $oldDisplay + ' -> ' + $newDisplay)
                $rowChanged = $true
            } elseif ($typeLabel.ToUpperInvariant() -eq 'ACTUAL') {
                $prefix = ''
                if (-not [string]::IsNullOrWhiteSpace($RunLabel)) {
                    $prefix = '[' + $RunLabel + '] '
                }
                $remark = Append-Remark $remark ($prefix + $carrierLabel + 'ACTUAL ARRIVAL_CONFIRMED: ' + $newDisplay)
            }
        }

        if (-not [string]::IsNullOrWhiteSpace($remark)) {
            $remarkCell = $sheet.Cells.Item($row, $trackingNoteCol)
            $existing = ([string]$remarkCell.Text).Trim()
            $remarkCell.Value2 = Merge-Tracking-Note $existing $remark
            if ($remark.StartsWith('ERROR') -or $remark.StartsWith('NOT_FOUND')) {
                $remarkCell.Interior.Color = 13421823
            } elseif ($remark.Contains('NO_ARRIVAL_DATE')) {
                $remarkCell.Interior.Color = 13495295
            }
            $rowChanged = $true
        }

        if ($rowChanged) {
            $updated++
        }
    }

    if ([IO.Path]::GetFullPath($WorkbookPath) -eq [IO.Path]::GetFullPath($OutputPath)) {
        $workbook.Save()
    } else {
        $workbook.SaveAs($OutputPath)
    }
    Write-Output "UPDATED_ROWS=$updated"
}
catch {
    $line = $_.InvocationInfo.ScriptLineNumber
    $message = $_.Exception.Message
    Write-Error ("Excel COM script failed at line {0}: {1}" -f $line, $message)
    exit 1
}
finally {
    if ($null -ne $workbook) {
        $workbook.Close($false)
        [Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }
    $excel.Quit()
    [Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
'''
