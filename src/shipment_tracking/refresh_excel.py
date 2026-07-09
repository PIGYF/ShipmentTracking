from __future__ import annotations

import argparse
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import socket
import sys
import time
from urllib.error import URLError

from openpyxl import load_workbook

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from shipment_tracking.dgf import DgfClient
from shipment_tracking.dsv import DsvClient
from shipment_tracking.env import load_dotenv
from shipment_tracking.excel_writer import (
    COL_BILL_NO,
    COL_EXCEPTION,
    COL_FORWARDER,
    COL_STATUS,
    COL_TRACKING_NOTE,
    update_tracking_workbook,
)
from shipment_tracking.maersk import MaerskClient
from shipment_tracking.models import TrackingRecord


DEFAULT_SHEET = "2026"
PENDING_STATUS = "\u672a\u9001\u8d27"
LOG_FILENAME = "refresh.log"
ENABLED_CARRIERS = {"DGF", "DSV", "MAERSK"}
_log_path: Path | None = None
_run_label: str = ""


@dataclass
class CarrierStats:
    queried: int = 0
    found: int = 0
    not_found: int = 0
    no_arrival_date: int = 0
    error: int = 0
    skipped: int = 0


def main() -> None:
    load_dotenv()
    parser = argparse.ArgumentParser(description="Refresh supported shipment tracking data in an Excel workbook.")
    parser.add_argument("excel_path", help="Input workbook path.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name. Default: 2026.")
    parser.add_argument("--output", help="Output workbook path. Default: overwrite input workbook.")
    parser.add_argument("--limit", type=int, help="Limit unique API queries for testing.")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be queried without calling APIs or writing Excel.")
    args = parser.parse_args()

    source = Path(args.excel_path)
    output = Path(args.output) if args.output else source
    if not source.exists():
        raise FileNotFoundError(f"Excel workbook not found: {source}")
    if args.output and not output.parent.exists():
        raise FileNotFoundError(f"Output directory not found: {output.parent}")
    _init_log(output)
    run_started = time.monotonic()
    _emit(f"Workbook: {source}")
    _emit(f"Sheet: {args.sheet}")
    _emit(f"Output: {output}")
    _emit(f"Enabled carriers: {', '.join(sorted(ENABLED_CARRIERS))}")
    if args.dry_run:
        _emit("Mode: DRY-RUN")
    rows = _load_pending_supported_rows(source, args.sheet)
    unique_jobs = _unique_jobs(rows)
    if args.limit:
        unique_jobs = unique_jobs[: args.limit]

    counts = Counter(carrier for carrier, _ in unique_jobs)
    _emit(f"Loaded pending supported rows: {len(rows)}")
    _emit("Unique API queries: " + ", ".join(f"{carrier}={count}" for carrier, count in sorted(counts.items())) if counts else "Unique API queries: 0")
    _emit_carrier_query_counts(counts)

    if args.dry_run:
        for index, (carrier, tracking_number) in enumerate(unique_jobs, start=1):
            _emit(f"[DRY-RUN] {index}/{len(unique_jobs)} {carrier} {tracking_number}")
        return

    records, remarks, stats = _run_jobs_by_carrier(unique_jobs)

    updated = update_tracking_workbook(source, records, output, args.sheet, remarks, run_label=_run_label)
    _emit_carrier_result_counts(stats)
    _emit(f"Updated rows: {updated}")
    _emit(f"Saved: {output}")
    _emit(f"Elapsed: {_format_elapsed(time.monotonic() - run_started)}")


def _run_jobs_by_carrier(unique_jobs: list[tuple[str, str]]) -> tuple[list[TrackingRecord], dict[str, str], dict[str, CarrierStats]]:
    grouped: dict[str, list[str]] = {carrier: [] for carrier in sorted(ENABLED_CARRIERS)}
    for carrier, tracking_number in unique_jobs:
        grouped.setdefault(carrier, []).append(tracking_number)

    active_groups = {carrier: jobs for carrier, jobs in grouped.items() if jobs}
    records: list[TrackingRecord] = []
    remarks: dict[str, str] = {}
    stats: dict[str, CarrierStats] = {}

    with ThreadPoolExecutor(max_workers=len(active_groups) or 1) as executor:
        futures = {
            executor.submit(_run_carrier_jobs, carrier, jobs): carrier
            for carrier, jobs in active_groups.items()
        }
        for future in as_completed(futures):
            carrier = futures[future]
            carrier_records, carrier_remarks, carrier_stats = future.result()
            records.extend(carrier_records)
            remarks.update(carrier_remarks)
            stats[carrier] = carrier_stats
            total = len(active_groups[carrier])
            _emit(f"[{carrier}] completed {total}/{total}")

    return records, remarks, stats


def _run_carrier_jobs(carrier: str, tracking_numbers: list[str]) -> tuple[list[TrackingRecord], dict[str, str], CarrierStats]:
    records: list[TrackingRecord] = []
    remarks: dict[str, str] = {}
    stats = CarrierStats(queried=len(tracking_numbers))
    total = len(tracking_numbers)
    try:
        client = _client_for(carrier)
    except RuntimeError as exc:
        skipped = _dated_remark(f"SKIPPED: {exc}")
        for tracking_number in tracking_numbers:
            remarks[tracking_number] = skipped
        stats.skipped = total
        _emit(f"[{carrier}] skipped {total}/{total}: {exc}")
        return records, remarks, stats

    for index, tracking_number in enumerate(tracking_numbers, start=1):
        try:
            record = _track(client, carrier, tracking_number)
            records.append(record)
            _update_stats(stats, record)
            remark = _remark(record)
            if remark:
                remarks[tracking_number] = _dated_remark(remark)
            _emit(_progress_line(index, total, record, remark))
        except Exception as exc:
            stats.error += 1
            error = _dated_remark(_format_error(exc))
            remarks[tracking_number] = error
            _emit(f"[{carrier}] {index}/{total} {tracking_number} -> {error}")
            if _should_stop_carrier(exc):
                skipped = tracking_numbers[index:]
                for skipped_tracking_number in skipped:
                    remarks[skipped_tracking_number] = _dated_remark(f"SKIPPED: {carrier} query stopped after connection/configuration error")
                stats.skipped += len(skipped)
                if skipped:
                    _emit(
                        f"[{carrier}] stopped after API connection/configuration error; "
                        f"skipped remaining {len(skipped)}/{total}"
                    )
                break

        if carrier == "DGF" and index < total:
            time.sleep(6)

    return records, remarks, stats


def _client_for(carrier: str):
    if carrier == "DGF":
        return DgfClient()
    if carrier == "DSV":
        return DsvClient()
    if carrier == "MAERSK":
        return MaerskClient()
    raise ValueError(f"Unsupported carrier: {carrier}")


def _load_pending_supported_rows(path: Path, sheet_name: str) -> list[tuple[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        header_row, headers = _find_header_row(sheet)
        columns = {header.lower(): index for index, header in enumerate(headers) if header}
        bill_idx = _find_column(columns, [COL_BILL_NO])
        carrier_idx = _find_column(columns, [COL_FORWARDER])
        status_idx = _find_column(columns, [COL_STATUS])
        tracking_note_idx = _find_optional_column(columns, [COL_TRACKING_NOTE])
        exception_idx = _find_optional_column(columns, [COL_EXCEPTION])

        rows: list[tuple[str, str]] = []
        skipped_actual_arrival = 0
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            status = str(row[status_idx] or "").strip()
            if status != PENDING_STATUS:
                continue
            carrier = str(row[carrier_idx] or "").strip().upper()
            if carrier not in ENABLED_CARRIERS:
                continue
            tracking_note = row[tracking_note_idx] if tracking_note_idx is not None else None
            exception = row[exception_idx] if exception_idx is not None else None
            if _has_actual_arrival_remark(carrier, tracking_note) or _has_actual_arrival_remark(carrier, exception):
                skipped_actual_arrival += 1
                continue
            tracking_number = str(row[bill_idx] or "").strip()
            if tracking_number:
                rows.append((carrier, tracking_number))
        if skipped_actual_arrival:
            _emit(f"Skipped rows with ACTUAL arrival remark: {skipped_actual_arrival}")
        return rows
    finally:
        workbook.close()


def _find_header_row(sheet) -> tuple[int, list[str]]:
    aliases = {COL_BILL_NO, COL_FORWARDER, COL_STATUS}
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        headers = [str(value).strip() if value is not None else "" for value in row]
        lowered = [header.lower() for header in headers]
        if all(alias.lower() in lowered for alias in aliases):
            return row_number, headers
    raise ValueError("Could not find expected header row with 提单号, 货代, 状态显示.")


def _unique_jobs(rows: list[tuple[str, str]]) -> list[tuple[str, str]]:
    seen: set[tuple[str, str]] = set()
    jobs: list[tuple[str, str]] = []
    for item in rows:
        if item in seen:
            continue
        seen.add(item)
        jobs.append(item)
    return jobs


def _find_column(columns: dict[str, int], names: list[str]) -> int:
    column = _find_optional_column(columns, names)
    if column is not None:
        return column
    raise ValueError(f"Missing required column: {', '.join(names)}")


def _find_optional_column(columns: dict[str, int], names: list[str]) -> int | None:
    for name in names:
        needle = name.lower()
        for header, index in columns.items():
            if header == needle or needle in header:
                return index
    return None


def _has_actual_arrival_remark(carrier: str, remark_value) -> bool:
    if remark_value is None:
        return False
    carrier = carrier.upper()
    for line in str(remark_value).splitlines():
        normalized = line.upper()
        if carrier in normalized and "ACTUAL" in normalized:
            return True
    return False


def _track(client, carrier: str, tracking_number: str) -> TrackingRecord:
    if carrier == "DGF":
        return client.track(tracking_number).to_record()
    if carrier == "DSV":
        return client.track(tracking_number).to_record()
    if carrier == "MAERSK":
        return client.track(tracking_number).to_record()
    raise ValueError(f"Unsupported carrier: {carrier}")


def _remark(record: TrackingRecord) -> str | None:
    if not record.found:
        return "NOT_FOUND: API returned no shipment"
    if not record.arrival_date:
        return "NO_ARRIVAL_DATE: shipment found but no ETA/ATA"
    return None


def _progress_line(index: int, total: int, record: TrackingRecord, remark: str | None) -> str:
    if remark:
        result = remark
    elif record.arrival_date:
        result = f"{record.arrival_date_type} {record.arrival_date:%Y-%m-%d %H:%M}"
    else:
        result = "OK"
    return f"[{record.carrier}] {index}/{total} {record.tracking_number} -> {result}"


def _format_error(exc: Exception) -> str:
    return f"ERROR: {type(exc).__name__}: {exc}"


def _update_stats(stats: CarrierStats, record: TrackingRecord) -> None:
    if not record.found:
        stats.not_found += 1
    else:
        stats.found += 1
        if not record.arrival_date:
            stats.no_arrival_date += 1


def _dated_remark(message: str) -> str:
    return f"[{_run_label}] {message}" if _run_label else message


def _emit_carrier_query_counts(counts: Counter[str]) -> None:
    _emit("Carrier query counts:")
    if not counts:
        _emit("  none")
        return
    for carrier in sorted(counts):
        _emit(f"  {carrier}: {counts[carrier]}")


def _emit_carrier_result_counts(stats: dict[str, CarrierStats]) -> None:
    _emit("Carrier results:")
    if not stats:
        _emit("  none")
        return
    for carrier in sorted(stats):
        item = stats[carrier]
        _emit(
            f"  {carrier}: queried={item.queried}, found={item.found}, "
            f"not_found={item.not_found}, no_arrival_date={item.no_arrival_date}, "
            f"error={item.error}, skipped={item.skipped}"
        )


def _format_elapsed(seconds: float) -> str:
    total = int(seconds)
    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def _should_stop_carrier(exc: Exception) -> bool:
    if isinstance(exc, RuntimeError):
        return True
    if isinstance(exc, (TimeoutError, ConnectionError, socket.timeout, URLError)):
        return True
    return False


def _init_log(excel_path: Path) -> None:
    global _log_path, _run_label
    _run_label = datetime.now().strftime("%Y-%m-%d %H:%M")
    _log_path = excel_path.parent / LOG_FILENAME
    _log_path.parent.mkdir(parents=True, exist_ok=True)
    _log_path.write_text(f"Run started: {_run_label}\n", encoding="utf-8")


def _emit(message: str) -> None:
    print(message, flush=True)
    if _log_path:
        with _log_path.open("a", encoding="utf-8") as log_file:
            log_file.write(f"{message}\n")


if __name__ == "__main__":
    main()
